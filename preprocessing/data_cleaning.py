"""Module for data cleaning functions."""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def change_value(column: int, previous_value: str, new_value: str, first_not_taken: str,
                 wb_path: str = './Data_sheet.xlsx'):
    """Change a value in a column to another value."""
    wb = load_workbook(wb_path)
    fill = PatternFill(start_color='FFFF00',
                       end_color='FFFF00', fill_type='solid')
    for sheet_name in wb.sheetnames[:-2]:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=2,
                                   max_row=2000,
                                   min_col=column,
                                   max_col=column):
            for cell in row:
                if cell.value == first_not_taken:
                    break
                if cell.value == previous_value:
                    cell.value = new_value
                    cell.fill = fill
            else:
                continue
            break
    wb.save('Updated_data_sheet.xlsx')


def find_unique_values(wb_path: str = './Data_sheet.xlsx'):
    """Find unique values for each column."""
    wb = load_workbook(wb_path)
    col_names = ['Line', 'Passage', 'Origin', 'Type', 'Subtype', 'Pools', 'CRISPR EDIT', 'Genotype',
                 'Reprogramming method', 'Age', 'Gender', 'Info', 'Media', 'ECM', 'Date',
                 'Initials', 'Notes', 'Cell number', 'Confluency']
    unique_values = [set() for _ in range(len(col_names))]
    for sheet_name in wb.sheetnames[:-2]:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=2,
                                   max_row=2000,
                                   min_col=2,
                                   max_col=len(col_names)+1,
                                   values_only=True):
            for (value_set, value, col_name) in zip(unique_values, row, col_names):
                if value == "Barcode/":  # Makes sure the last sheet's last part is cut out
                    break
                if value != col_name:
                    value_set.add(value)
            else:
                continue
            break

    for value_set in unique_values:
        value_set.remove(None)

    unique_values = [[f'{str(value)}\n' for value in unique_set]
                     for unique_set in unique_values]

    with open('unique_values.txt', 'w', encoding='utf8') as f:
        for (col, col_name) in zip(unique_values, col_names):
            f.write(f'{col_name}:\n')
            f.writelines(col)
            f.write('\n')


if __name__ == '__main__':
    # find_unique_values()
    change_value(13, "General Stock", "General stock", "5' barcode")
